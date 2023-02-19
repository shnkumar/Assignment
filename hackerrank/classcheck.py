import openpyxl as xl
import os
wb  = xl.load_workbook(r'G:\Python\TEST.xlsx')
print('sheet name : ', wb.sheetnames)
ws=wb['EMP']
print('sheet name : ', ws['B2'].value)

id_to_deleted=2
for row_num in range(2,ws.max_row+1):
    if ws.cell(row_num,1).value==id_to_deleted:
        ws.delete_rows(row_num,1)
        wb.save(r'G:\Python\TEST.xlsx')
        print('sheet name : ', ws['B2'].value,ws['B3'].value,row_num)
0