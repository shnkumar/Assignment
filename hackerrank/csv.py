import openpyxl as xl
import os
wb  = xl.load_workbook(r'G:\Python\TEST.xlsx')
print('sheet name : ', wb.sheetnames)
ws=wb['EMP']
cell1=ws.cell(1,1)
cell2=ws['A2']

print('cell 1 value', cell1.value)
print('cell 2 value', cell2.value)
print('cell 1 max row ', ws.max_row,ws.max_column)
cells=ws['A1':'C4']
for row in cells:
    for cell in row:
        print(cell.value,end=' ')
    print()
cells1=ws[1]
print('Item of row 1 are :',end=' ')
for cell in cells1:
    print(cell.value,end=' ')
cells1 =ws['C']
print('\nItems of column C are :')
for cell in cells1:
    print(cell.value)


